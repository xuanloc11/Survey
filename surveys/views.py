from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout, get_user_model
from django.contrib.auth.hashers import make_password, check_password
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.core.files.storage import FileSystemStorage
from django.core.mail import send_mail
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.tokens import default_token_generator
from django.conf import settings
from django.urls import reverse
from django.core import signing
from uuid import uuid4
import json
import csv
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Survey, Question, Response
from .forms import (
    SurveyForm,
    QuestionForm,
    ResponseForm,
    UserRegisterForm,
    UserProfileForm,
)

User = get_user_model()


def register_view(request):
    if request.user.is_authenticated:
        return redirect('surveys:home')

    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user: User = form.save(commit=False)
            user.is_active = False
            user.save()

            uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            activation_link = request.build_absolute_uri(
                reverse('surveys:activate', kwargs={'uidb64': uidb64, 'token': token})
            )

            subject = 'Xác nhận tài khoản HCMUTE Survey'
            message = (
                f'Xin chào {user.username},\n\n'
                'Cảm ơn bạn đã đăng ký tài khoản trên HCMUTE Survey.\n'
                'Vui lòng nhấp vào liên kết dưới đây để kích hoạt tài khoản của bạn:\n\n'
                f'{activation_link}\n\n'
                'Nếu bạn không thực hiện đăng ký này, hãy bỏ qua email.\n\n'
                'Trân trọng,\n'
                'HCMUTE Survey'
            )

            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
                messages.success(
                    request,
                    'Đăng ký tài khoảnthành công! Vui lòng kiểm tra email để xác nhận tài khoản trước khi đăng nhập.'
                )
            except Exception:
                messages.warning(
                    request,
                    'Đăng ký thành công nhưng hiện không gửi được email xác nhận. '
                    'Hãy liên hệ quản trị viên để kích hoạt tài khoản.'
                )

            return redirect('surveys:login')
    else:
        form = UserRegisterForm()

    return render(request, 'auth/register.html', {'form': form})


def activate_account(request, uidb64, token):
    """Kích hoạt tài khoản từ link trong email"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        if not user.is_active:
            user.is_active = True
            user.save()
        messages.success(request, 'Tài khoản của bạn đã được kích hoạt. Bây giờ bạn có thể đăng nhập.')
        return redirect('surveys:login')

    messages.error(request, 'Link kích hoạt không hợp lệ hoặc đã hết hạn.')
    return redirect('surveys:login')


def password_reset_request(request):
    """Form nhập email để nhận link đặt lại mật khẩu"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            messages.error(request, 'Vui lòng nhập email đã đăng ký.')
        else:
            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                user = None

            if user is None:
                messages.error(request, 'Không tìm thấy tài khoản nào với email này.')
            else:
                uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
                token = default_token_generator.make_token(user)
                reset_link = request.build_absolute_uri(
                    reverse('surveys:password_reset_confirm', kwargs={'uidb64': uidb64, 'token': token})
                )

                subject = 'Đặt lại mật khẩu tài khoản hệ thống khảo sát HCMUTE'
                message = (
                    f'Xin chào {user.username},\n\n'
                    'Bạn vừa yêu cầu đặt lại mật khẩu cho tài khoản HCMUTE Survey.\n'
                    'Vui lòng nhấp vào liên kết dưới đây để đặt mật khẩu mới:\n\n'
                    f'{reset_link}\n\n'
                    'Nếu bạn không yêu cầu, hãy bỏ qua email này.\n\n'
                    'Trân trọng,\n'
                    'Đội ngũ HCMUTE Survey'
                )

                try:
                    send_mail(
                        subject=subject,
                        message=message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[user.email],
                        fail_silently=False,
                    )
                    messages.success(
                        request,
                        'Đã gửi email hướng dẫn đặt lại mật khẩu. Vui lòng kiểm tra hộp thư của bạn.'
                    )
                    return redirect('surveys:login')
                except Exception:
                    messages.error(
                        request,
                        'Hiện không gửi được email đặt lại mật khẩu. Vui lòng thử lại sau hoặc liên hệ quản trị viên.'
                    )

    return render(request, 'auth/password_reset_request.html')


def password_reset_confirm(request, uidb64, token):
    """Đặt mật khẩu mới từ link trong email"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        messages.error(request, 'Link đặt lại mật khẩu không hợp lệ hoặc đã hết hạn.')
        return redirect('surveys:login')

    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not password1 or not password2:
            messages.error(request, 'Vui lòng nhập đầy đủ mật khẩu mới.')
        elif password1 != password2:
            messages.error(request, 'Mật khẩu nhập lại không khớp.')
        elif len(password1) < 6:
            messages.error(request, 'Mật khẩu phải có ít nhất 6 ký tự.')
        else:
            user.set_password(password1)
            user.save()
            messages.success(request, 'Đã đặt lại mật khẩu thành công. Bây giờ bạn có thể đăng nhập.')
            return redirect('surveys:login')

    return render(request, 'auth/password_reset_confirm.html', {'uidb64': uidb64, 'token': token})


def login_view(request):
    """Trang đăng nhập"""
    if request.user.is_authenticated:
        return redirect('surveys:survey_list')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('surveys:survey_list')
        else:
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng!')
    
    return render(request, 'auth/login.html')


def logout_view(request):
    """Trang đăng xuất"""
    logout(request)
    messages.success(request, 'Bạn đã đăng xuất thành công!')
    return redirect('surveys:home')


@login_required
def profile_view(request):
    """Trang thông tin cá nhân"""
    from .models import UserProfile
    UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật thông tin cá nhân!')
            return redirect('surveys:profile')
    else:
        form = UserProfileForm(instance=request.user)

    return render(request, 'auth/profile.html', {'form': form})


def home(request):
    """Landing page chính, giới thiệu sản phẩm trước khi vào khảo sát"""
    context = {
        'total_surveys': Survey.objects.filter(is_active=True).count(),
        'total_responses': Response.objects.count(),
    }
    return render(request, 'surveys/pages/home.html', context)
@login_required
def dashboard(request):
    """Trang dashboard cho người dùng đã đăng nhập"""
    total_surveys = Survey.objects.count()
    total_responses = Response.objects.count()

    user_surveys = Survey.objects.filter(creator=request.user).annotate(
        response_count=Count('responses')
    ).order_by('-created_at')
    user_surveys_count = user_surveys.count()
    user_responses_count = Response.objects.filter(survey__creator=request.user).count()

    context = {
        'total_surveys': total_surveys,
        'total_responses': total_responses,
        'user_surveys_count': user_surveys_count,
        'user_responses_count': user_responses_count,
        'recent_user_surveys': user_surveys[:5],
    }
    return render(request, 'surveys/dashboard/dashboard.html', context)


@login_required
def survey_list(request):
    """Danh sách khảo sát của người dùng"""
    surveys = Survey.objects.filter(creator=request.user).annotate(
        response_count=Count('responses')
    ).order_by('-created_at')
    
    context = {
        'surveys': surveys,
    }
    return render(request, 'surveys/survey/survey_list.html', context)

@login_required
def survey_create(request):
    """Tạo khảo sát mới"""
    if request.method == 'POST':
        form = SurveyForm(request.POST, request.FILES)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.creator = request.user
            raw_password = form.cleaned_data.get('password')
            if raw_password:
                survey.password = make_password(raw_password)
            survey.is_active = False
            survey.save()
            messages.success(request, 'Đã tạo khảo sát ở trạng thái nháp. Hãy xuất bản khi sẵn sàng!')
            return redirect('surveys:survey_detail', pk=survey.pk)
    else:
        form = SurveyForm(initial={'is_active': False})
    
    return render(request, 'surveys/survey/survey_form.html', {
        'form': form,
        'title': 'Tạo khảo sát mới'
    })


@login_required
def survey_edit(request, pk):
    """Chỉnh sửa khảo sát"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    
    if request.method == 'POST':
        form = SurveyForm(request.POST, request.FILES, instance=survey)
        if form.is_valid():
            survey = form.save(commit=False)
            raw_password = form.cleaned_data.get('password')
            if raw_password:
                survey.password = make_password(raw_password)
            survey.save()
            messages.success(request, 'Đã cập nhật khảo sát thành công!')
            return redirect('surveys:survey_detail', pk=survey.pk)
    else:
        form = SurveyForm(instance=survey)
    
    return render(request, 'surveys/survey/survey_form.html', {
        'form': form,
        'survey': survey,
        'title': 'Chỉnh sửa khảo sát'
    })


def survey_detail(request, pk):
    """Chi tiết khảo sát"""
    survey = get_object_or_404(Survey, pk=pk)
    questions = survey.questions.all().order_by('order')
    
    is_expired = survey.expires_at and survey.expires_at < timezone.now()
    responses_count = survey.responses.count()
    max_responses = survey.max_responses
    is_limit_reached = bool(max_responses) and responses_count >= max_responses
    remaining_slots = max_responses - responses_count if max_responses else None
    
    can_edit = request.user.is_authenticated and request.user == survey.creator
    template_name = 'surveys/survey/survey_builder.html' if can_edit else 'surveys/survey/survey_detail.html'
    
    context = {
        'survey': survey,
        'questions': questions,
        'is_expired': is_expired,
        'is_limit_reached': is_limit_reached,
        'remaining_slots': remaining_slots,
        'can_edit': can_edit,
        'builder_mode': can_edit,
        'share_link': request.build_absolute_uri(
            reverse('surveys:survey_take_token', args=[signing.dumps(survey.pk, salt="survey-share")])
        ),
    }
    
    if can_edit:
        responses = survey.responses.all()
        total_responses_count = responses.count()
        
        stats = []
        for question in questions:
            question_id_str = str(question.id)
            
            if question.question_type not in ['text', 'single', 'multiple']:
                continue
            
            if question.question_type == 'text':
                text_answers = []
                for response in responses:
                    if response.response_data and question_id_str in response.response_data:
                        answer_value = response.response_data[question_id_str]
                        if isinstance(answer_value, str) and answer_value.strip():
                            text_answers.append(answer_value)
                
                stats.append({
                    'question': question,
                    'type': 'text',
                    'answers': text_answers[:10],  # Hiển thị 10 câu trả lời đầu
                    'total': len(text_answers)
                })
            else:
                choice_stats = []
                if question.options:
                    for idx, option_text in enumerate(question.options):
                        count = 0
                        for response in responses:
                            if response.response_data and question_id_str in response.response_data:
                                answer_value = response.response_data[question_id_str]
                                if question.question_type == 'single':
                                    # So sánh với option text
                                    if answer_value == option_text:
                                        count += 1
                                elif question.question_type == 'multiple':
                                    # Kiểm tra nếu option có trong danh sách
                                    if isinstance(answer_value, list) and option_text in answer_value:
                                        count += 1
                        
                        percentage = (count / total_responses_count * 100) if total_responses_count > 0 else 0
                        choice_stats.append({
                            'option': option_text,
                            'index': idx,
                            'count': count,
                            'percentage': round(percentage, 1)
                        })
                
                stats.append({
                    'question': question,
                    'type': question.question_type,
                    'choices': choice_stats,
                    'total': total_responses_count
                })
        
        context['stats'] = stats
        context['total_responses'] = total_responses_count
        context['responses'] = responses
    
    return render(request, template_name, context)


@login_required
def survey_delete(request, pk):
    """Xóa khảo sát"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    
    if request.method == 'POST':
        survey.delete()
        messages.success(request, 'Đã xóa khảo sát thành công!')
        return redirect('surveys:survey_list')
    
    return render(request, 'surveys/survey/survey_confirm_delete.html', {
        'survey': survey
    })


@login_required
def question_add(request, survey_pk):
    """Thêm câu hỏi vào khảo sát"""
    survey = get_object_or_404(Survey, pk=survey_pk, creator=request.user)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.survey = survey
            question.save()
            
            if question.question_type in ['single', 'multiple']:
                return redirect('surveys:choice_add', question_pk=question.pk)
            else:
                messages.success(request, 'Đã thêm câu hỏi thành công!')
                return redirect('surveys:survey_detail', pk=survey.pk)
    else:
        form = QuestionForm()
    
    return render(request, 'surveys/question/question_form.html', {
        'form': form,
        'survey': survey,
        'title': 'Thêm câu hỏi'
    })


@login_required
def question_edit(request, pk):
    """Chỉnh sửa câu hỏi"""
    question = get_object_or_404(Question, pk=pk)
    survey = question.survey
    
    if survey.creator != request.user:
        messages.error(request, 'Bạn không có quyền chỉnh sửa câu hỏi này!')
        return redirect('surveys:survey_detail', pk=survey.pk)
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã cập nhật câu hỏi thành công!')
            return redirect('surveys:survey_detail', pk=survey.pk)
    else:
        form = QuestionForm(instance=question)
    
    return render(request, 'surveys/question/question_form.html', {
        'form': form,
        'survey': survey,
        'question': question,
        'title': 'Chỉnh sửa câu hỏi'
    })


@login_required
def question_delete(request, pk):
    """Xóa câu hỏi"""
    question = get_object_or_404(Question, pk=pk)
    survey = question.survey
    
    if survey.creator != request.user:
        messages.error(request, 'Bạn không có quyền xóa câu hỏi này!')
        return redirect('surveys:survey_detail', pk=survey.pk)
    
    if request.method == 'POST':
        survey_pk = survey.pk
        question.delete()
        messages.success(request, 'Đã xóa câu hỏi thành công!')
        return redirect('surveys:survey_detail', pk=survey_pk)
    
    return render(request, 'surveys/question/question_confirm_delete.html', {
        'question': question,
        'survey': survey
    })


@login_required
def choice_add(request, question_pk):
    """Thêm lựa chọn cho câu hỏi"""
    question = get_object_or_404(Question, pk=question_pk)
    survey = question.survey
    
    if survey.creator != request.user:
        messages.error(request, 'Bạn không có quyền thêm lựa chọn!')
        return redirect('surveys:survey_detail', pk=survey.pk)
    
    if request.method == 'POST':
        option_text = request.POST.get('text', '').strip()
        if option_text:
            if question.options is None:
                question.options = []
            question.options.append(option_text)
            question.save()
            
            if 'add_another' in request.POST:
                messages.success(request, 'Đã thêm lựa chọn! Tiếp tục thêm...')
                return redirect('surveys:choice_add', question_pk=question.pk)
            else:
                messages.success(request, 'Đã thêm lựa chọn thành công!')
                return redirect('surveys:survey_detail', pk=survey.pk)
        else:
            messages.error(request, 'Vui lòng nhập nội dung lựa chọn!')
    
    options = question.options or []
    
    return render(request, 'surveys/choice/choice_form.html', {
        'question': question,
        'survey': survey,
        'options': options,
        'title': 'Thêm lựa chọn'
    })


def survey_take(request, pk):
    """Trang làm khảo sát"""
    survey = get_object_or_404(Survey, pk=pk, is_active=True)
    session_key = f'survey_access_{survey.id}'
    done_session_key = f'survey_done_{survey.id}'

    back_url = request.META.get('HTTP_REFERER')
    current_url = request.build_absolute_uri(request.get_full_path())
    if back_url:
        if back_url.rstrip('/') == current_url.rstrip('/'):
            back_url = None
    if not back_url:
        back_url = reverse('surveys:survey_detail', args=[survey.pk])

    if not request.session.session_key:
        request.session.create()
    if 'anon_session_id' not in request.session:
        request.session['anon_session_id'] = uuid4().hex
    
    whitelist_raw = survey.whitelist_emails or ""
    whitelist = {email.strip().lower() for email in whitelist_raw.splitlines() if email.strip()}

    responses_count = survey.responses.count()
    if survey.max_responses and responses_count >= survey.max_responses:
        messages.error(request, 'Khảo sát đã đạt tới giới hạn số phản hồi.')
        return redirect('surveys:survey_detail', pk=pk)
    if survey.password:
        has_access = request.session.get(session_key)
        if not has_access:
            if request.method == 'POST' and 'survey_password' in request.POST:
                entered_password = request.POST.get('survey_password', '')
                if entered_password and check_password(entered_password, survey.password):
                    request.session[session_key] = True
                    messages.success(request, 'Đã xác nhận mật khẩu khảo sát.')
                    return redirect('surveys:survey_take', pk=pk)
                else:
                    messages.error(request, 'Mật khẩu không đúng. Vui lòng thử lại.')
            return render(request, 'surveys/survey/survey_take.html', {
                'survey': survey,
                'questions': [],
                'need_password': True,
                'back_url': back_url,
            })
    
    if survey.expires_at and survey.expires_at < timezone.now():
        messages.error(request, 'Khảo sát này đã hết hạn!')
        return redirect('surveys:survey_detail', pk=pk)
    
    if whitelist and request.user.is_authenticated:
        user_email = (request.user.email or '').lower()
        if user_email not in whitelist and request.user != survey.creator:
            messages.error(request, 'Email của bạn không nằm trong whitelist tham gia khảo sát.')
            return redirect('surveys:survey_detail', pk=pk)

    if request.user.is_authenticated:
        existing_response = Response.objects.filter(survey=survey, respondent=request.user).first()
        if existing_response:
            if survey.creator == request.user:
                messages.info(request, 'Bạn đã tham gia khảo sát này rồi!')
                return redirect('surveys:survey_results', pk=pk)
            else:
                if not survey.one_response_only:
                    pass
                else:
                    if survey.allow_review_response:
                        messages.info(request, 'Bạn đã tham gia khảo sát này rồi! Đang chuyển đến xem lại câu trả lời của bạn.')
                        return redirect('surveys:survey_review_response', response_id=existing_response.id)
                    elif not survey.allow_review_response and survey.send_confirmation_email:
                        messages.info(request, 'Bạn đã tham gia khảo sát này rồi!')
                        return redirect('surveys:survey_thankyou', pk=pk)
                    else:
                        messages.info(request, 'Bạn đã tham gia khảo sát này rồi!')
                        return redirect('surveys:survey_detail', pk=pk)
    else:
        if request.session.get(done_session_key):
            # Kiểm tra xem có response_id trong session không
            response_id_key = f'survey_response_{survey.id}'
            response_id = request.session.get(response_id_key)
            
            # Kiểm tra xem có cho phép trả lời nhiều lần không
            if not survey.one_response_only:
                # Cho phép trả lời nhiều lần → Không chặn, cho làm lại
                pass
            else:
                if response_id:
                    # Redirect dựa vào cài đặt
                    if survey.allow_review_response:
                        messages.info(request, 'Bạn đã tham gia khảo sát này rồi! Đang chuyển đến xem lại câu trả lời của bạn.')
                        return redirect('surveys:survey_review_response', response_id=response_id)
                    elif not survey.allow_review_response and survey.send_confirmation_email:
                        messages.info(request, 'Bạn đã tham gia khảo sát này rồi!')
                        return redirect('surveys:survey_thankyou', pk=pk)
                
                messages.info(request, 'Bạn đã tham gia khảo sát này rồi từ thiết bị này!')
                return redirect('surveys:survey_detail', pk=pk)
            
        client_ip = get_client_ip(request)
        existing_response_ip = Response.objects.filter(
            survey=survey,
            respondent__isnull=True,
            ip_address=client_ip
        ).first()
        if existing_response_ip:
            # Kiểm tra xem có cho phép trả lời nhiều lần không
            if not survey.one_response_only:
                # Cho phép trả lời nhiều lần → Không chặn, cho làm lại
                pass
            else:
                # Chỉ cho trả lời 1 lần → Redirect dựa vào cài đặt
                if survey.allow_review_response:
                    messages.info(request, 'Bạn đã tham gia khảo sát này rồi! Đang chuyển đến xem lại câu trả lời của bạn.')
                    return redirect('surveys:survey_review_response', response_id=existing_response_ip.id)
                elif not survey.allow_review_response and survey.send_confirmation_email:
                    messages.info(request, 'Bạn đã tham gia khảo sát này rồi!')
                    return redirect('surveys:survey_thankyou', pk=pk)
                else:
                    messages.info(request, 'Bạn đã tham gia khảo sát này rồi từ thiết bị này!')
                    return redirect('surveys:survey_detail', pk=pk)
    
    if request.method == 'POST':
        # Xác minh Cloudflare Turnstile Captcha cho người dùng ẩn danh
        if not request.user.is_authenticated:
            cf_response = request.POST.get('cf-turnstile-response')
            
            if not cf_response:
                messages.error(request, 'Vui lòng hoàn thành xác minh captcha.')
                questions = survey.questions.all()
                return render(request, 'surveys/survey/survey_take.html', {
                    'survey': survey,
                    'questions': questions,
                    'need_password': False,
                    'back_url': back_url,
                    'TURNSTILE_SITE_KEY': settings.CLOUDFLARE_TURNSTILE_SITE_KEY,
                })
            
            # Verify với Cloudflare
            verify_url = 'https://challenges.cloudflare.com/turnstile/v0/siteverify'
            verify_data = {
                'secret': settings.CLOUDFLARE_TURNSTILE_SECRET_KEY,
                'response': cf_response,
                'remoteip': get_client_ip(request)
            }
            
            try:
                verify_result = requests.post(verify_url, data=verify_data, timeout=10).json()
                
                if not verify_result.get('success'):
                    error_codes = verify_result.get('error-codes', [])
                    messages.error(request, 'Xác minh captcha thất bại. Vui lòng thử lại.')
                    questions = survey.questions.all()
                    return render(request, 'surveys/survey/survey_take.html', {
                        'survey': survey,
                        'questions': questions,
                        'need_password': False,
                        'back_url': back_url,
                        'TURNSTILE_SITE_KEY': settings.CLOUDFLARE_TURNSTILE_SITE_KEY,
                    })
            except requests.RequestException as e:
                messages.error(request, 'Không thể xác minh captcha. Vui lòng thử lại sau.')
                questions = survey.questions.all()
                return render(request, 'surveys/survey/survey_take.html', {
                    'survey': survey,
                    'questions': questions,
                    'need_password': False,
                    'back_url': back_url,
                    'TURNSTILE_SITE_KEY': settings.CLOUDFLARE_TURNSTILE_SITE_KEY,
                })
        
        errors = []
        for question in survey.questions.all():
            field_name = f'question_{question.id}'
            if question.question_type not in ['text', 'single', 'multiple']:
                continue
            if question.is_required:
                if question.question_type == 'multiple':
                    if field_name not in request.POST or not request.POST.getlist(field_name):
                        errors.append(f'Vui lòng trả lời câu hỏi: {question.text}')
                else:
                    if field_name not in request.POST or not request.POST.get(field_name):
                        errors.append(f'Vui lòng trả lời câu hỏi: {question.text}')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            response_data = {}
            for question in survey.questions.all():
                field_name = f'question_{question.id}'
                
                if question.question_type == 'text':
                    text_answer = request.POST.get(field_name, '').strip()
                    if text_answer:
                        response_data[str(question.id)] = text_answer
                elif question.question_type == 'single':
                    selected_index = request.POST.get(field_name)
                    if selected_index and question.options:
                        try:
                            index = int(selected_index)
                            if 0 <= index < len(question.options):
                                response_data[str(question.id)] = question.options[index]
                        except (ValueError, IndexError):
                            pass
                elif question.question_type == 'multiple':
                    selected_indices = request.POST.getlist(field_name)
                    if selected_indices and question.options:
                        selected_options = []
                        for idx_str in selected_indices:
                            try:
                                index = int(idx_str)
                                if 0 <= index < len(question.options):
                                    selected_options.append(question.options[index])
                            except (ValueError, IndexError):
                                pass
                        if selected_options:
                            response_data[str(question.id)] = selected_options
            
            response = Response.objects.create(
                survey=survey,
                respondent=request.user if request.user.is_authenticated else None,
                ip_address=get_client_ip(request),
                response_data=response_data
            )
            request.session[done_session_key] = True
            # Lưu response_id vào session để có thể redirect lại sau này
            response_id_key = f'survey_response_{survey.id}'
            request.session[response_id_key] = response.id
            
            messages.success(request, 'Cảm ơn bạn đã tham gia khảo sát!')
            
            # Logic xử lý dựa vào cài đặt
            if survey.allow_review_response:
                # Bật xem lại câu trả lời → redirect đến trang review
                return redirect('surveys:survey_review_response', response_id=response.id)
            else:
                # Tắt xem lại câu trả lời
                if survey.send_confirmation_email and request.user.is_authenticated and request.user.email:
                    # Gửi email xác nhận nếu có bật
                    try:
                        from django.template.loader import render_to_string
                        from django.utils import timezone
                        
                        email_context = {
                            'survey': survey,
                            'user_email': request.user.email,
                            'completion_time': timezone.now(),
                        }
                        
                        html_message = render_to_string('surveys/email/survey_confirmation.html', email_context)
                        
                        send_mail(
                            subject=f'Cảm ơn bạn đã tham gia: {survey.title}',
                            message=f'Cảm ơn bạn đã hoàn thành khảo sát "{survey.title}". Câu trả lời của bạn đã được ghi nhận.',
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[request.user.email],
                            html_message=html_message,
                            fail_silently=True,
                        )
                    except Exception as e:
                        # Log error nhưng không làm gián đoạn flow
                        pass
                    
                    # Redirect đến trang thank you
                    return redirect('surveys:survey_thankyou', pk=pk)
                else:
                    # Không gửi email → redirect về survey detail
                    return redirect('surveys:survey_detail', pk=pk)
    
    questions = survey.questions.all()
    
    return render(request, 'surveys/survey/survey_take.html', {
        'survey': survey,
        'questions': questions,
        'need_password': False,
        'back_url': back_url,
        'TURNSTILE_SITE_KEY': settings.CLOUDFLARE_TURNSTILE_SITE_KEY,
    })

def survey_take_token(request, token):
    """Truy cập khảo sát bằng link token hóa"""
    try:
        pk = signing.loads(token, salt="survey-share", max_age=None)
    except signing.BadSignature:
        messages.error(request, 'Link khảo sát không hợp lệ hoặc đã bị thay đổi.')
        return redirect('surveys:home')
    return redirect('surveys:survey_take', pk=pk)

def survey_review_response(request, response_id):
    """Xem lại câu trả lời đã submit"""
    response = get_object_or_404(Response, pk=response_id)
    survey = response.survey
    
    # Kiểm tra xem khảo sát có cho phép xem lại câu trả lời không
    if not survey.allow_review_response:
        # Chỉ creator mới được xem
        if request.user != survey.creator:
            messages.error(request, 'Khảo sát này không cho phép xem lại câu trả lời.')
            return redirect('surveys:survey_detail', pk=survey.pk)
    
    # Kiểm tra quyền xem: chỉ người tạo response hoặc creator của survey
    if request.user.is_authenticated:
        if response.respondent != request.user and survey.creator != request.user:
            messages.error(request, 'Bạn không có quyền xem câu trả lời này.')
            return redirect('surveys:home')
    else:
        # Người dùng ẩn danh: kiểm tra session hoặc IP
        client_ip = get_client_ip(request)
        if response.respondent is not None or response.ip_address != client_ip:
            messages.error(request, 'Bạn không có quyền xem câu trả lời này.')
            return redirect('surveys:home')
    
    # Lấy danh sách câu hỏi và câu trả lời
    questions = survey.questions.all()
    response_data = response.response_data or {}
    
    # Tạo dict chứa câu hỏi và câu trả lời
    questions_with_answers = []
    for question in questions:
        if question.question_type in ['text', 'single', 'multiple']:
            answer = response_data.get(str(question.id))
            questions_with_answers.append({
                'question': question,
                'answer': answer
            })
    
    return render(request, 'surveys/survey/survey_review.html', {
        'survey': survey,
        'response': response,
        'questions_with_answers': questions_with_answers,
    })

def survey_thankyou(request, pk):
    """Trang cảm ơn sau khi hoàn thành khảo sát"""
    survey = get_object_or_404(Survey, pk=pk, is_active=True)
    
    # Lấy email và thời gian hoàn thành
    user_email = None
    if request.user.is_authenticated:
        user_email = request.user.email
    
    return render(request, 'surveys/survey/survey_thankyou.html', {
        'survey': survey,
        'user_email': user_email,
        'completion_time': timezone.now(),
    })

def custom_404(request, exception=None):
    return render(request, 'errors/404.html', status=404)


def custom_500(request):
    return render(request, 'errors/500.html', status=500)


def custom_502(request):
    return render(request, 'errors/502.html', status=502)


def custom_404_preview(request):
    """Preview 404 page while DEBUG=True"""
    return render(request, 'errors/404.html', status=404)


@login_required
def survey_results(request, pk):
    """Kết quả khảo sát"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    
    responses = survey.responses.all()
    questions = survey.questions.all().order_by('order')
    
    stats = []
    total_responses_count = responses.count()
    
    for question in questions:
        question_id_str = str(question.id)
        
        if question.question_type not in ['text', 'single', 'multiple']:
            continue
        
        if question.question_type == 'text':
            text_answers = []
            for response in responses:
                if response.response_data and question_id_str in response.response_data:
                    answer_value = response.response_data[question_id_str]
                    if isinstance(answer_value, str) and answer_value.strip():
                        text_answers.append(answer_value)
            
            stats.append({
                'question': question,
                'type': 'text',
                'answers': text_answers[:10],  # Hiển thị 10 câu trả lời đầu
                'total': len(text_answers)
            })
        else:
            choice_stats = []
            if question.options:
                for idx, option_text in enumerate(question.options):
                    count = 0
                    for response in responses:
                        if response.response_data and question_id_str in response.response_data:
                            answer_value = response.response_data[question_id_str]
                            if question.question_type == 'single':
                                if answer_value == option_text:
                                    count += 1
                            elif question.question_type == 'multiple':
                                if isinstance(answer_value, list) and option_text in answer_value:
                                    count += 1
                    
                    percentage = (count / total_responses_count * 100) if total_responses_count > 0 else 0
                    choice_stats.append({
                        'option': option_text,
                        'index': idx,
                        'count': count,
                        'percentage': round(percentage, 1)
                    })
            
            stats.append({
                'question': question,
                'type': question.question_type,
                'choices': choice_stats,
                'total': total_responses_count
            })
    
    context = {
        'survey': survey,
        'responses': responses,
        'stats': stats,
        'total_responses': total_responses_count
    }
    return render(request, 'surveys/survey/survey_results.html', context)


@login_required
def survey_export_csv(request, pk):
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)

    responses = survey.responses.all().order_by('submitted_at')
    questions = survey.questions.all().order_by('order')

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="khao_sat_{survey.pk}_ket_qua.csv"'
    response.write("\ufeff")  # BOM for Excel compatibility
    writer = csv.writer(response, delimiter=',')
    header = ['Thời gian']
    for question in questions:
        header.append(question.text)
    writer.writerow(header)

    for resp in responses:
        submitted_local = timezone.localtime(resp.submitted_at)
        time_display = submitted_local.strftime("%d/%m/%Y %H:%M:%S")
        row = [time_display]

        for question in questions:
            qid = str(question.id)
            answer = ''
            if resp.response_data and qid in resp.response_data:
                value = resp.response_data[qid]
                if isinstance(value, list):
                    answer = ' | '.join(str(v) for v in value)
                else:
                    answer = str(value)
            row.append(answer)

        writer.writerow(row)

    return response


@login_required
def survey_export_excel(request, pk):
    """Xuất kết quả khảo sát ra file Excel với định dạng đẹp"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    
    responses = survey.responses.all().order_by('submitted_at')
    questions = survey.questions.all().order_by('order')
    
    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Khảo sát {survey.pk}"
    
    # Định dạng cho header
    header_fill = PatternFill(start_color="5B2C6F", end_color="5B2C6F", fill_type="solid")  # Màu tím
    header_font = Font(bold=True, color="FFFFFF", size=12)  # Chữ trắng, đậm
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Tạo header row
    headers = ['Thời gian']
    for question in questions:
        headers.append(question.text)
    
    # Ghi header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Định dạng cho data rows
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Màu xám nhạt
    
    # Ghi dữ liệu
    for row_num, resp in enumerate(responses, 2):
        submitted_local = timezone.localtime(resp.submitted_at)
        time_display = submitted_local.strftime("%d/%m/%Y %H:%M:%S")
        
        # Cột thời gian
        cell = ws.cell(row=row_num, column=1)
        cell.value = time_display
        cell.alignment = data_alignment
        cell.border = border
        if row_num % 2 == 0:  # Dòng chẵn có màu nền
            cell.fill = alternate_fill
        
        # Các cột câu trả lời
        for col_num, question in enumerate(questions, 2):
            qid = str(question.id)
            answer = ''
            if resp.response_data and qid in resp.response_data:
                value = resp.response_data[qid]
                if isinstance(value, list):
                    answer = ', '.join(str(v) for v in value)
                else:
                    answer = str(value)
            
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = answer
            cell.alignment = data_alignment
            cell.border = border
            if row_num % 2 == 0:  # Dòng chẵn có màu nền
                cell.fill = alternate_fill
    
    # Tự động điều chỉnh độ rộng cột
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        # Tính độ dài tối đa của cột
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Đặt độ rộng (tối đa 50, tối thiểu 15)
        adjusted_width = min(max(max_length + 2, 15), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Đóng băng hàng đầu tiên
    ws.freeze_panes = 'A2'
    
    # Tạo HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="khao_sat_{survey.pk}_ket_qua.xlsx"'
    
    # Lưu workbook vào response
    wb.save(response)
    
    return response


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@login_required
@require_http_methods(["POST"])
def question_add_ajax(request, survey_pk):
    """API AJAX để thêm câu hỏi"""
    survey = get_object_or_404(Survey, pk=survey_pk, creator=request.user)
    
    try:
        data = json.loads(request.body)
        question = Question.objects.create(
            survey=survey,
            text=data.get('text', ''),
            question_type=data.get('question_type', 'single'),
            order=data.get('order', survey.questions.count() + 1),
            is_required=data.get('is_required', True),
            subtitle=data.get('subtitle', ''),
            media_url=data.get('media_url', '')
        )
        
        options_data = data.get('choices', []) or data.get('options', [])
        question.options = [opt.strip() for opt in options_data if opt and opt.strip()]
        question.save()
        
        return JsonResponse({
            'success': True,
            'question': {
                'id': question.id,
                'text': question.text,
            'question_type': question.question_type,
            'is_required': question.is_required,
            'order': question.order,
            'options': question.options or [],
            'subtitle': question.subtitle,
            'media_url': question.media_url
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST"])
def question_update_ajax(request, pk):
    question = get_object_or_404(Question, pk=pk)
    
    if question.survey.creator != request.user:
        return JsonResponse({'success': False, 'error': 'Không có quyền'}, status=403)
    
    try:
        data = json.loads(request.body)
        question.text = data.get('text', question.text)
        question.question_type = data.get('question_type', question.question_type)
        question.order = data.get('order', question.order)
        question.is_required = data.get('is_required', question.is_required)
        question.subtitle = data.get('subtitle', question.subtitle)
        question.media_url = data.get('media_url', question.media_url)
        
        if 'choices' in data or 'options' in data:
            options_data = data.get('choices', []) or data.get('options', [])
            question.options = [opt.strip() for opt in options_data if opt and opt.strip()]

        if 'correct_answers' in data:
            raw_correct = data.get('correct_answers') or []
            cleaned = []
            for idx in raw_correct:
                try:
                    i = int(idx)
                    cleaned.append(i)
                except (TypeError, ValueError):
                    continue
            question.correct_answers = cleaned

        question.save()
        
        return JsonResponse({
            'success': True,
            'question': {
                'id': question.id,
                'text': question.text,
            'question_type': question.question_type,
            'is_required': question.is_required,
            'order': question.order,
            'options': question.options or [],
            'correct_answers': question.correct_answers or [],
            'subtitle': question.subtitle,
            'media_url': question.media_url
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def question_delete_ajax(request, pk):
    """API AJAX để xóa câu hỏi"""
    question = get_object_or_404(Question, pk=pk)
    
    if question.survey.creator != request.user:
        return JsonResponse({'success': False, 'error': 'Không có quyền'}, status=403)
    
    try:
        question.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def question_reorder_ajax(request, survey_pk):
    """API AJAX để sắp xếp lại thứ tự câu hỏi"""
    survey = get_object_or_404(Survey, pk=survey_pk, creator=request.user)
    
    try:
        data = json.loads(request.body)
        question_orders = data.get('orders', [])  # [{id: 1, order: 1}, {id: 2, order: 2}]
        
        for item in question_orders:
            Question.objects.filter(pk=item['id'], survey=survey).update(order=item['order'])
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def survey_publish_toggle_ajax(request, pk):
    """Bật/tắt xuất bản khảo sát"""
    survey = get_object_or_404(Survey, pk=pk, creator=request.user)
    try:
        data = json.loads(request.body or "{}")
        is_active = bool(data.get('is_active', True))
        survey.is_active = is_active
        survey.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': survey.is_active})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def question_image_upload_ajax(request, pk):
    """Upload ảnh cho câu hỏi và lưu đường dẫn vào media_url"""
    question = get_object_or_404(Question, pk=pk)

    if question.survey.creator != request.user:
        return JsonResponse({'success': False, 'error': 'Không có quyền'}, status=403)

    image_file = request.FILES.get('image')
    if not image_file:
        return JsonResponse({'success': False, 'error': 'Không có file ảnh'}, status=400)

    try:
        fs = FileSystemStorage(
            location=settings.MEDIA_ROOT / 'question_images',
            base_url=settings.MEDIA_URL + 'question_images/'
        )
        filename = fs.save(image_file.name, image_file)
        file_url = fs.url(filename)

        question.media_url = file_url
        question.save(update_fields=['media_url'])

        return JsonResponse({'success': True, 'url': file_url})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def choice_add_ajax(request, question_pk):
    """API AJAX để thêm lựa chọn cho câu hỏi"""
    question = get_object_or_404(Question, pk=question_pk)
    
    if question.survey.creator != request.user:
        return JsonResponse({'success': False, 'error': 'Không có quyền'}, status=403)
    
    try:
        data = json.loads(request.body)
        option_text = data.get('text', '').strip()
        
        if not option_text:
            return JsonResponse({'success': False, 'error': 'Vui lòng nhập nội dung lựa chọn!'}, status=400)
        
        # Thêm vào options (JSONField)
        if question.options is None:
            question.options = []
        question.options.append(option_text)
        question.save()
        
        return JsonResponse({
            'success': True,
            'choice': {
                'text': option_text,
                'index': len(question.options) - 1
            },
            'options': question.options
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def choice_delete_ajax(request, pk):
    """API AJAX để xóa lựa chọn khỏi câu hỏi"""
    question = get_object_or_404(Question, pk=pk)
    
    if question.survey.creator != request.user:
        return JsonResponse({'success': False, 'error': 'Không có quyền'}, status=403)
    
    try:
        data = json.loads(request.body)
        choice_index = data.get('index')
        
        if choice_index is None:
            return JsonResponse({'success': False, 'error': 'Thiếu thông tin index'}, status=400)
        
        if question.options is None or not isinstance(question.options, list):
            return JsonResponse({'success': False, 'error': 'Không có lựa chọn nào'}, status=400)
        
        try:
            index = int(choice_index)
            if 0 <= index < len(question.options):
                removed_option = question.options.pop(index)
                question.save()
                return JsonResponse({
                    'success': True,
                    'removed_option': removed_option,
                    'options': question.options
                })
            else:
                return JsonResponse({'success': False, 'error': 'Index không hợp lệ'}, status=400)
        except (ValueError, IndexError):
            return JsonResponse({'success': False, 'error': 'Index không hợp lệ'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


