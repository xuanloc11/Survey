import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from ..models import Survey, ResponseAttachment
from ..permissions import get_survey_access


@login_required
def survey_results(request, pk):
    survey = get_object_or_404(Survey, pk=pk, is_deleted=False)
    access = get_survey_access(request.user, survey)
    if not access.can_view_results:
        return render(request, 'errors/404.html', status=404)

    responses = survey.responses.all()
    questions = survey.questions.all().order_by('order')

    stats = []
    total_responses_count = responses.count()

    for question in questions:
        question_id_str = str(question.id)

        if question.question_type not in ['text', 'single', 'multiple', 'upload']:
            continue

        if question.question_type == 'text':
            text_answers = []
            for response in responses:
                if response.response_data and question_id_str in response.response_data:
                    answer_value = response.response_data[question_id_str]
                    if isinstance(answer_value, str) and answer_value.strip():
                        text_answers.append(answer_value)

            stats.append({
                'question': question,
                'type': 'text',
                'answers': text_answers[:10],  # Hiển thị 10 câu trả lời đầu
                'total': len(text_answers)
            })
        elif question.question_type == 'upload':
            # Count & sample attachments
            q_attachments = (
                ResponseAttachment.objects
                .filter(response__survey=survey, question=question)
                .select_related("response")
                .order_by("-uploaded_at")
            )
            total = q_attachments.count()
            samples = list(q_attachments[:10])
            stats.append({
                'question': question,
                'type': 'upload',
                'attachments': samples,
                'total': total,
            })
        else:
            choice_stats = []
            if question.options:
                for idx, option_text in enumerate(question.options):
                    count = 0
                    for response in responses:
                        if response.response_data and question_id_str in response.response_data:
                            answer_value = response.response_data[question_id_str]
                            if question.question_type == 'single':
                                if answer_value == option_text:
                                    count += 1
                            elif question.question_type == 'multiple':
                                if isinstance(answer_value, list) and option_text in answer_value:
                                    count += 1

                    percentage = (count / total_responses_count * 100) if total_responses_count > 0 else 0
                    choice_stats.append({
                        'option': option_text,
                        'index': idx,
                        'count': count,
                        'percentage': round(percentage, 1)
                    })

            stats.append({
                'question': question,
                'type': question.question_type,
                'choices': choice_stats,
                'total': total_responses_count
            })

    context = {
        'survey': survey,
        'responses': responses,
        'stats': stats,
        'total_responses': total_responses_count
    }
    return render(request, 'surveys/survey_management/survey_results.html', context)


@login_required
def survey_export_csv(request, pk):
    survey = get_object_or_404(Survey, pk=pk, is_deleted=False)
    access = get_survey_access(request.user, survey)
    if not access.can_view_results:
        return render(request, 'errors/404.html', status=404)

    responses = survey.responses.all().order_by('submitted_at')
    questions = survey.questions.all().order_by('order')

    attachments = (
        ResponseAttachment.objects
        .filter(response__survey=survey)
        .select_related("response", "question")
    )
    attachment_map = {(a.response_id, a.question_id): a for a in attachments}

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="khao_sat_{survey.pk}_ket_qua.csv"'
    response.write("\ufeff")  # BOM for Excel compatibility
    writer = csv.writer(response, delimiter=',')
    header = ['Thời gian']
    for question in questions:
        header.append(question.text)
    writer.writerow(header)

    for resp in responses:
        submitted_local = timezone.localtime(resp.submitted_at)
        time_display = submitted_local.strftime("%d/%m/%Y %H:%M:%S")
        row = [time_display]

        for question in questions:
            qid = str(question.id)
            answer = ''
            if question.question_type == 'upload':
                att = attachment_map.get((resp.id, question.id))
                if att:
                    # export absolute URL for convenience
                    answer = request.build_absolute_uri(att.file.url)
            else:
                if resp.response_data and qid in resp.response_data:
                    value = resp.response_data[qid]
                    if isinstance(value, list):
                        answer = ' | '.join(str(v) for v in value)
                    else:
                        answer = str(value)
            row.append(answer)

        writer.writerow(row)

    return response


@login_required
def survey_export_excel(request, pk):
    survey = get_object_or_404(Survey, pk=pk, is_deleted=False)
    access = get_survey_access(request.user, survey)
    if not access.can_view_results:
        return render(request, 'errors/404.html', status=404)

    responses = survey.responses.all().order_by('submitted_at')
    questions = survey.questions.all().order_by('order')

    attachments = (
        ResponseAttachment.objects
        .filter(response__survey=survey)
        .select_related("response", "question")
    )
    attachment_map = {(a.response_id, a.question_id): a for a in attachments}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Khảo sát {survey.pk}"

    header_fill = PatternFill(start_color="0023ff", end_color="0023ff", fill_type="solid")  # Màu tím
    header_font = Font(bold=True, color="FFFFFF", size=12)  # Chữ trắng, đậm
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    headers = ['Thời gian']
    for question in questions:
        headers.append(question.text)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Màu xám nhạt

    for row_num, resp in enumerate(responses, 2):
        submitted_local = timezone.localtime(resp.submitted_at)
        time_display = submitted_local.strftime("%d/%m/%Y %H:%M:%S")

        cell = ws.cell(row=row_num, column=1)
        cell.value = time_display
        cell.alignment = data_alignment
        cell.border = border
        if row_num % 2 == 0:  # Dòng chẵn có màu nền
            cell.fill = alternate_fill

        for col_num, question in enumerate(questions, 2):
            qid = str(question.id)
            answer = ''
            if question.question_type == 'upload':
                att = attachment_map.get((resp.id, question.id))
                if att:
                    answer = request.build_absolute_uri(att.file.url)
            else:
                if resp.response_data and qid in resp.response_data:
                    value = resp.response_data[qid]
                    if isinstance(value, list):
                        answer = ', '.join(str(v) for v in value)
                    else:
                        answer = str(value)

            cell = ws.cell(row=row_num, column=col_num)
            cell.value = answer
            cell.alignment = data_alignment
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = alternate_fill

    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = min(max(max_length + 2, 15), 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="khao_sat_{survey.pk}_ket_qua.xlsx"'
    wb.save(response)
    return response


